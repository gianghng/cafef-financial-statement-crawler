import json
import argparse
import asyncio
from typing import List

import httpx
import lxml.html
from openpyxl import Workbook
from loguru import logger

ascii_chars: str = 'aaaaaaaaaaaaaaaaaeeeeeeeeeeediiiiiooooooooooooooooouuuuuuuuuuuyyyyyAAAAAAAAAAAAAAAAAEEEEEEEEEEEDIIIOOOOOOOOOOOOOOOOOOOUUUUUUUUUUUYYYYYAADOOU'
uni_chars: str = 'àáảãạâầấẩẫậăằắẳẵặèéẻẽẹêềếểễệđìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵÀÁẢÃẠÂẦẤẨẪẬĂẰẮẲẴẶÈÉẺẼẸÊỀẾỂỄỆĐÌÍỈĨỊÒÓỎÕỌÔỒỐỔỖỘƠỜỚỞỠỢÙÚỦŨỤƯỪỨỬỮỰỲÝỶỸỴÂĂĐÔƠƯ'
company: dict = {}
headers: dict = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
    'Accept-Encoding': 'gzip, deflate',
    'Host': 's.cafef.vn',
}


def get_company():
    global company

    if not len(company):
        with open('company.json') as f:
            company = json.load(f)

    return company


def unicode_to_ascii(raw_str: str) -> str:
    ret_val: str = ''
    raw_str = raw_str.strip()
    s: str = raw_str

    arr_ascii: str = ascii_chars

    for i in range(len(s)):
        pos: int = uni_chars.find(s[i])
        if pos >= 0:
            ret_val += arr_ascii[pos]
        else:
            ret_val += s[i]

    return ret_val


def unicode_to_ascii_with_score(raw_str: str) -> str:
    str_char: str = "abcdefghiklmnopqrstxyzuvxw0123456789 "
    raw_str = raw_str.strip()

    _str: str = raw_str.replace("–", "")
    _str = _str.replace("  ", " ")
    _str = unicode_to_ascii(_str.lower())

    s: str = _str
    s_return: str = ""

    for i in range(len(s)):
        if str_char.find(s[i]) > -1:
            if s[i] != ' ':
                s_return += s[i]
            elif i > 0 and s[i - 1] != ' ' and s[i - 1] != '-':
                s_return += '-'

    return s_return


def symbol_to_slug(symbol: str) -> str:
    return unicode_to_ascii_with_score(get_company()[symbol])


def build_url(symbol: str, report_type: str, year: int, quarter: int, show_type: int, idx: int):
    return f"https://s.cafef.vn/bao-cao-tai-chinh" \
           f"/{symbol}/{report_type}/{year}/{quarter}/{idx}/{show_type}" \
           f"/bao-cao-tai-chinh-{symbol_to_slug(symbol)}.chn"


async def crawl_financial_report_by_type(symbol: str, report_type: str, year: int, quarter: int, show_type: int, idx: int) -> list:
    async with httpx.AsyncClient(http2=True) as client:
        r = await client.get(
            url=build_url(
                symbol, report_type, year, quarter, show_type, idx,
            ),
            headers=headers,
        )

    tree = lxml.html.fromstring(r.text)

    element_table_headers = tree.xpath("//table[@id='tblGridData']")[0]
    element_table = tree.xpath("//table[@id='tableContent']")[0]
    elements_table_tr: list = element_table.xpath("tr")

    excel_rows = [["#", *[e.text.strip() for e in element_table_headers.xpath("tr[1]//td")[1:-1]]]]

    for element_table_tr in elements_table_tr:
        excel_row = []
        for element_table_td in element_table_tr.xpath("td")[:-1]:
            _str = "".join(element_table_td.itertext()).strip()
            excel_row.append(_str)

        excel_rows.append(excel_row)

    return excel_rows


async def save_to_excel(wb: Workbook, sheet_name: str, data_rows: List[List[str]]):
    wb.create_sheet(sheet_name)

    sheet = wb[sheet_name]

    for row_index, cols_data in enumerate(data_rows, start=1):
        for col_index, val in enumerate(cols_data, start=1):
            sheet.cell(row=row_index, column=col_index).value = val


async def crawl_financial_report(symbol: str, year: int, quarter: int, show_type: int, idx: int):
    report_types = ['BSheet', 'IncSta', 'CashFlow', 'CashFlowDirect']

    wb = Workbook()

    for report_type in report_types:
        data = await crawl_financial_report_by_type(symbol, report_type, year, quarter, show_type, idx)

        await save_to_excel(wb, report_type, data)

    wb.save(f"data/{symbol}-{year}-{quarter}.xlsx")


async def read_symbols(file_path: str) -> List:
    symbols: list = []
    with open(file_path) as f:
        for row in f:
            data: str = row.strip()
            symbols.append(data.split(","))
    return symbols


async def main():
    parser = argparse.ArgumentParser(description='Lấy dữ liệu báo cáo tài chính từ cafef')
    parser.add_argument('file', help='Đường dẫn tới file CSV')
    args = parser.parse_args()

    symbols = await read_symbols(args.file)

    for symbol in symbols:
        logger.info(symbol)
        await crawl_financial_report(symbol[0], symbol[1], symbol[2], 0, 0)
        logger.success(symbol)
        await asyncio.sleep(1)


asyncio.run(main())
